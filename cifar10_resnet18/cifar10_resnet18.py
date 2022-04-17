from openpyxl import Workbook
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import copy
import numpy as np
np.set_printoptions(threshold=np.inf)
from torchvision import datasets, transforms
import torch
import math

from utils.sampling import mnist_iid, mnist_noniid, cifar_iid
from utils.options import args_parser
from models.Update import LocalUpdate
from models.Nets import resnet18
from models.Fed import FedAvg
from models.mask import *
from models.test import test_img

import time
start=time.time()

args = args_parser()

seed=1
np.random.seed(seed)
torch.manual_seed(seed)
torch.cuda.manual_seed_all(seed)
torch.backends.cudnn.deterministic = True
torch.backends.cudnn.benchmark = False


if __name__ == '__main__':    
    args.device = torch.device('cuda:{}'.format(args.gpu) if torch.cuda.is_available() and args.gpu != -1 else 'cpu')        
    # load dataset and split users        
    if args.dataset == 'cifar10':        
        transform1 = transforms.Compose([transforms.RandomHorizontalFlip(),transforms.RandomGrayscale(),transforms.ToTensor(),transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))])
        transform2 = transforms.Compose([transforms.ToTensor(), transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))])
        dataset_train = datasets.CIFAR10('./data/cifar', train=True, download=True, transform=transform1)
        dataset_test = datasets.CIFAR10('./data/cifar', train=False, download=True, transform=transform2)
        if args.iid:
            dict_users = cifar_iid(dataset_train, args.num_users)
        else:
            #exit('Error: only consider IID setting in CIFAR10')
            dict_users = cifar_noniid(dataset_train, args.num_users)
    else:
        exit('Error: unrecognized dataset')
    img_size = dataset_train[0][0].shape   
#model-------------------------------------------------------
    if args.model == 'lenet5' and args.dataset == 'fashionmnist':
        net_glob = lenet5(args=args).to(args.device)
    elif args.model == 'vanillacnn' and args.dataset == 'fashionmnist':
        net_glob = vanillacnn(args=args).to(args.device)
    elif args.model == 'resnet18' and args.dataset == 'cifar10':
        net_glob = resnet18().to(args.device)
    else:
        exit('Error: unrecognized model') 
    net_glob.train()
    
    parm1={}
    for name,parameters in net_glob.named_parameters():
        print(name,':',parameters.size())
    
    glob_acc, glob_train_loss = [],[]
    wb = Workbook()
    ws = wb.active
    
    resolution=0.005
    avg_losss=[]
    bitsvolume=0
    for epochs in range(args.epochs):        
        wglob=copy.deepcopy(net_glob.state_dict())
        tmp=copy.deepcopy(net_glob.state_dict())
        m = max(int(args.frac * args.num_users), 1)
        idxs_users = np.random.choice(range(args.num_users), m, replace=False)       
        dw_locals=[]; loss_locals = []
        
        bitsround=0
        for idx in idxs_users:         
            local = LocalUpdate(args=args, dataset=dataset_train, idxs=dict_users[idx])
            w, mode, loss = local.train(net=copy.deepcopy(net_glob).to(args.device))
            bitsnode=0
            for k in w.keys():                
                dw=copy.deepcopy(w[k])-wglob[k]
                bit=max(torch.ceil(torch.log2(rang(dw)/resolution)),2)
                dwquant=quandequan_range(dw,bit)                 
                tmp[k].data.copy_(dwquant)
                
                if 'batches_tracked' in k:
                    bitsnode+=0
                else:
                    bitsnode+=bit*w[k].numel()
                
            bitsround+=bitsnode
            dw_locals.append(copy.deepcopy(tmp))
            loss_locals.append(copy.deepcopy(loss))
        dw_glob=FedAvg(dw_locals)
        
        for k in wglob.keys():
            wglob[k]+=dw_glob[k]
        net_glob.load_state_dict(wglob)
        
        net_glob.to(args.device)
        net_glob.eval()        
        test_acc, test_loss = test_img(net_glob, dataset_test, args)
        train_acc, train_loss = test_img(net_glob, dataset_train, args) 
        avg_loss = sum(loss_locals) / len(loss_locals)
        avg_losss.append(avg_loss)
         
        print(test_acc.numpy())
        print('Avg_loss:',avg_loss)
        print('Train_loss {:.3f}, Test_loss {:.3f}'.format(train_loss, test_loss))
        
        bitsvolume+=bitsround
        
        ws.cell(epochs+2,1).value = epochs  
        ws.cell(epochs+2,2).value = str(test_acc.numpy()) 
        ws.cell(epochs+2,3).value = str(test_loss)
        ws.cell(epochs+2,4).value = str(train_acc.numpy())
        ws.cell(epochs+2,5).value = str(train_loss)
        ws.cell(epochs+2,6).value = str(avg_loss)
        ws.cell(epochs+2,7).value = str((bitsvolume/(1024**3)).cpu().numpy())
        ws.cell(epochs+2,8).value = str(bit)
                                    
    wb.save("./result/cifar10_resnet18.xlsx")
    
end=time.time()
print('running time is ',end-start)
