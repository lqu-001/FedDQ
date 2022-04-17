from openpyxl import Workbook
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import copy
import numpy as np
np.set_printoptions(threshold=np.inf)
from torchvision import datasets, transforms
import torch
import math

from utils.sampling import fashionmnist_iid, fashionmnist_noniid

from utils.options import args_parser
from models.Update import LocalUpdate
from models.Nets import vanillacnn
from models.Fed import FedAvg
from models.mask import mask,quandequan,rang
from models.test import test_img

args = args_parser() 
np.random.seed(1)

def setup_seed(seed):
     torch.manual_seed(seed)
     torch.cuda.manual_seed_all(seed)
     torch.backends.cudnn.deterministic = True
setup_seed(1)

if __name__ == '__main__':    
    args.device = torch.device('cuda:{}'.format(args.gpu) if torch.cuda.is_available() and args.gpu != -1 else 'cpu')
# load dataset and split users-------------------------------        
    if args.dataset == 'fashionmnist': 
        transform = transforms.Compose([transforms.ToTensor(), transforms.Normalize((0.5,), (0.5,))]) 
        dataset_train = datasets.FashionMNIST('./data/FashionMNIST', download=True, train=True, transform=transform)
        dataset_test = datasets.FashionMNIST('./data/FashionMNIST', download=True, train=False, transform=transform)
        if args.iid:
            dict_users = fashionmnist_iid(dataset_train, args.num_users)
        else:
            dict_users = fashionmnist_noniid(dataset_train, args.num_users)
    else:
        exit('Error: unrecognized dataset')
    img_size = dataset_train[0][0].shape    
#model-------------------------------------------------------
    if args.model == 'lenet5' and args.dataset == 'fashionmnist':
        net_glob = lenet5(args=args).to(args.device)
    elif args.model == 'vanillacnn' and args.dataset == 'fashionmnist':
        net_glob = vanillacnn(args=args).to(args.device)
    else:
        exit('Error: unrecognized model')
    print(net_glob)
    
    parm1={}
    for name,parameters in net_glob.named_parameters():
        print(name,':',parameters.size())
    
    glob_acc, glob_train_loss = [],[]
    wb = Workbook()
    ws = wb.active  
    resolution=0.005
    bitsvolume=0
    for iter in range(args.epochs):
        m = max(int(args.frac * args.num_users), 1)
        idxs_users = np.random.choice(range(args.num_users), m, replace=False)       
        w_locals, loss_locals = [], []
        conv1dws=[]; conv2dws=[]; conv3dws=[]; conv4dws=[]; fc1dws=[];  fc2dws=[];fc3dws=[]
        bitcv1s=[]; bitcv2s=[]; bitcv3s=[]; bitcv4s=[]; bitfc1s=[]; bitfc2s=[]; bitfc3s=[]
        bitround=0
        for idx in idxs_users: 
            agg_mode=net_glob
            local = LocalUpdate(args=args, dataset=dataset_train, idxs=dict_users[idx])
            w, mode, loss = local.train(net=copy.deepcopy(net_glob).to(args.device))                                     
            crt_mode=mode
            
            crtparm={}
            for name,parameters in crt_mode.named_parameters():
                crtparm[name]=parameters.detach().cpu().numpy()                
            aggparm={}
            for name,parameters in agg_mode.named_parameters():
                aggparm[name]=parameters.detach().cpu().numpy()
              
            conv1dw=crtparm['conv1.weight']-aggparm['conv1.weight']
            conv2dw=crtparm['conv2.weight']-aggparm['conv2.weight']
            fc1dw=crtparm['fc1.weight']-aggparm['fc1.weight']
            fc2dw=crtparm['fc2.weight']-aggparm['fc2.weight']
            
            bitcv1=math.ceil(math.log(rang(conv1dw)/resolution,2))
            bitcv2=math.ceil(math.log(rang(conv2dw)/resolution,2))
            bitfc1=math.ceil(math.log(rang(fc1dw)/resolution,2))
            bitfc2=math.ceil(math.log(rang(fc2dw)/resolution,2))
            
            conv1dw_quan=quandequan(conv1dw,bitcv1)
            conv2dw_quan=quandequan(conv2dw,bitcv2)
            fc1dw_quan=quandequan(fc1dw,bitfc1)
            fc2dw_quan=quandequan(fc2dw,bitfc2)
            conv1dw=conv1dw_quan.cpu().detach().numpy()
            conv2dw=conv2dw_quan.cpu().detach().numpy()
            fc1dw=fc1dw_quan.cpu().detach().numpy()
            fc2dw=fc2dw_quan.cpu().detach().numpy() 
            
            bitcv1s.append(bitcv1)
            bitcv2s.append(bitcv2)
            bitfc1s.append(bitfc1)
            bitfc2s.append(bitfc2)
            
            bitnode=bitcv1*32*1*5*5 + bitcv2*64*32*5*5 + bitfc1*512*1024 + bitfc2*10*512
            bitround=bitround+bitnode                
#-----server-------- 
            conv1dws.append(conv1dw)
            conv2dws.append(conv2dw)
            fc1dws.append(fc1dw)
            fc2dws.append(fc2dw)

            loss_locals.append(copy.deepcopy(loss))
         
        conv1dw_avg= np.mean(conv1dws, axis=0) 
        conv2dw_avg= np.mean(conv2dws, axis=0) 
        fc1dw_avg= np.mean(fc1dws, axis=0)
        fc2dw_avg= np.mean(fc2dws, axis=0)
        
        glob={}
        for name,parameters in net_glob.named_parameters():
            glob[name]=parameters.detach().cpu().numpy()
        
        conv1w=glob['conv1.weight'] + conv1dw_avg
        conv2w=glob['conv2.weight'] + conv2dw_avg 
        fc1w=glob['fc1.weight'] + fc1dw_avg
        fc2w=glob['fc2.weight'] + fc2dw_avg
                
        net_glob.conv1.weight.data=torch.from_numpy(conv1w)  
        net_glob.conv2.weight.data=torch.from_numpy(conv2w)  
        net_glob.fc1.weight.data=torch.from_numpy(fc1w) 
        net_glob.fc2.weight.data=torch.from_numpy(fc2w)
                   
        net_glob.to(args.device)
        net_glob.eval()
        test_acc, test_loss = test_img(net_glob, dataset_test, args)
        train_acc, train_loss = test_img(net_glob, dataset_train, args) 
        avg_loss = sum(loss_locals) / len(loss_locals)

        print(test_acc.numpy())
        bitsvolume=bitsvolume+bitround
        avgbit=bitsvolume/((iter+1)*10*(32*1*5*5 + 64*32*5*5 + 512*1024 + 10*512))
        
        cv1avgbit= np.mean(bitcv1s)
        cv2avgbit= np.mean(bitcv2s)
        fc1avgbit= np.mean(bitfc1s)
        fc2avgbit= np.mean(bitfc2s)
        
        ws.cell(iter+2,1).value = iter  
        ws.cell(iter+2,2).value = str(test_acc.numpy()) 
        ws.cell(iter+2,3).value = str(test_loss)
        ws.cell(iter+2,4).value = str(train_acc.numpy())
        ws.cell(iter+2,5).value = str(train_loss)
        ws.cell(iter+2,6).value = str(avg_loss)
        ws.cell(iter+2,7).value = bitsvolume
        ws.cell(iter+2,8).value = avgbit
        
        ws.cell(iter+2,9).value = cv1avgbit
        ws.cell(iter+2,10).value = cv2avgbit
        ws.cell(iter+2,11).value = fc1avgbit
        ws.cell(iter+2,12).value = fc2avgbit
                            
    wb.save("./result/fashion_mnist.xlsx")
